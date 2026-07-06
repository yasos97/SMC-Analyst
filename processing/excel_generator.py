"""
SalamaIQ - Generateur de Rapports Excel
=========================================
Genere un classeur Excel professionnel avec :
  - Feuille Resume (KPIs + graphiques)
  - Feuille Evolution Mensuelle (tableau + graphique)
  - Feuille Top 10 Clients
  - Feuille Detail Transactions
"""

import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint, SeriesLabel
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter
from datetime import datetime


# ── Couleurs SalamaIQ ──────────────────────────────────────────────────────────
SALAMA_RED = 'E8321A'
DARK_BLUE = '0F172A'
SLATE_600 = '475569'
LIGHT_BG = 'F8FAFC'
WHITE = 'FFFFFF'
BORDER_COLOR = 'E2E8F0'

# Styles reusables
HEADER_FONT = Font(name='Calibri', bold=True, size=11, color=WHITE)
HEADER_FILL = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type='solid')
TITLE_FONT = Font(name='Calibri', bold=True, size=16, color=DARK_BLUE)
SUBTITLE_FONT = Font(name='Calibri', bold=True, size=12, color=SALAMA_RED)
KPI_VALUE_FONT = Font(name='Calibri', bold=True, size=14, color=DARK_BLUE)
KPI_LABEL_FONT = Font(name='Calibri', size=10, color=SLATE_600)
NORMAL_FONT = Font(name='Calibri', size=10)
THIN_BORDER = Border(
    left=Side(style='thin', color=BORDER_COLOR),
    right=Side(style='thin', color=BORDER_COLOR),
    top=Side(style='thin', color=BORDER_COLOR),
    bottom=Side(style='thin', color=BORDER_COLOR)
)


def _format_number(val, decimals=2):
    """Formate un nombre pour l'affichage."""
    if val is None:
        return 0
    return round(float(val), decimals)


def _apply_table_style(ws, start_row, end_row, start_col, end_col):
    """Applique un style tableau propre a une plage."""
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if row == start_row:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
            else:
                cell.font = NORMAL_FONT
                if row % 2 == 0:
                    cell.fill = PatternFill(start_color=LIGHT_BG, end_color=LIGHT_BG, fill_type='solid')


def generate_excel_report(data_kpis, df_transactions, client_name="Tous les clients", period="N/A", compare=True):
    """
    Genere un rapport Excel complet avec graphiques.
    
    Args:
        data_kpis: dict des KPIs (depuis compute_kpis)
        df_transactions: DataFrame filtre
        client_name: Nom du client
        period: Periode d'analyse
        compare: Si True, inclut les donnees N-1
    
    Returns:
        bytes du fichier Excel
    """
    wb = Workbook()
    
    # ─────────────────────────────────────────────────────────────────────────
    # FEUILLE 1 : RESUME
    # ─────────────────────────────────────────────────────────────────────────
    ws_resume = wb.active
    ws_resume.title = "Resume"
    ws_resume.sheet_properties.tabColor = SALAMA_RED
    
    # Titre
    ws_resume.merge_cells('A1:H1')
    cell_title = ws_resume['A1']
    titre = "RELEVE DE CONSOMMATION CLIENT" if compare else "ANALYSE D'ACTIVITE"
    cell_title.value = f"SalamaIQ - {titre}"
    cell_title.font = TITLE_FONT
    cell_title.alignment = Alignment(horizontal='center', vertical='center')
    ws_resume.row_dimensions[1].height = 35
    
    # Info client / periode
    ws_resume.merge_cells('A2:H2')
    ws_resume['A2'].value = f"Client : {client_name}  |  Periode : {period}  |  Genere le {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws_resume['A2'].font = Font(name='Calibri', size=10, color=SLATE_600)
    ws_resume['A2'].alignment = Alignment(horizontal='center')
    ws_resume.row_dimensions[2].height = 25
    
    # ── KPIs ──
    ws_resume.merge_cells('A4:H4')
    ws_resume['A4'].value = "INDICATEURS CLES"
    ws_resume['A4'].font = SUBTITLE_FONT
    
    current_year = data_kpis.get('current_year', 'N')
    previous_year = data_kpis.get('previous_year', 'N-1')
    
    # En-tetes KPI
    row = 6
    if compare:
        headers = ['Indicateur', f'Valeur {current_year}', f'Valeur {previous_year}', 'Variation (%)', 'Unite']
    else:
        headers = ['Indicateur', f'Valeur {current_year}', 'Unite']
    
    for col_idx, h in enumerate(headers, 1):
        cell = ws_resume.cell(row=row, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER
    
    # Donnees KPI
    kpi_rows = [
        ("Chiffre d'Affaires HT", 'ca_total', 'ca_total_n1', 'ca_variation', 'MAD'),
        ("Marge HT", 'marge_ht', 'marge_ht_n1', 'marge_variation', 'MAD'),
        ("Volume Total", 'volume_total', 'volume_total_n1', 'vol_total_variation', 'L'),
        ("Volume Gasoil", 'volume_gasoil', 'volume_gasoil_n1', 'vol_gasoil_variation', 'L'),
        ("Volume Super SP", 'volume_super', 'volume_super_n1', 'vol_super_variation', 'L'),
        ("Taux de Marge", 'taux_marge', 'taux_marge_n1', None, '%'),
        ("Marge Unitaire", 'marge_unitaire', 'marge_unitaire_n1', 'marge_unitaire_variation', 'MAD/L'),
    ]
    
    for i, (label, key_n, key_n1, key_var, unit) in enumerate(kpi_rows):
        r = row + 1 + i
        val_n = _format_number(data_kpis.get(key_n, 0))
        
        if compare:
            val_n1 = _format_number(data_kpis.get(key_n1, 0))
            variation = data_kpis.get(key_var) if key_var else None
            
            ws_resume.cell(row=r, column=1, value=label).font = Font(name='Calibri', bold=True, size=10)
            ws_resume.cell(row=r, column=2, value=val_n).number_format = '#,##0.00'
            ws_resume.cell(row=r, column=3, value=val_n1).number_format = '#,##0.00'
            
            var_cell = ws_resume.cell(row=r, column=4)
            if variation is not None:
                var_cell.value = variation / 100
                var_cell.number_format = '+0.0%;-0.0%'
                var_cell.font = Font(name='Calibri', bold=True, size=10, 
                                     color='16A34A' if variation >= 0 else 'DC2626')
            else:
                var_cell.value = "N/A"
            
            ws_resume.cell(row=r, column=5, value=unit)
        else:
            ws_resume.cell(row=r, column=1, value=label).font = Font(name='Calibri', bold=True, size=10)
            ws_resume.cell(row=r, column=2, value=val_n).number_format = '#,##0.00'
            ws_resume.cell(row=r, column=3, value=unit)
        
        # Bordures
        for c in range(1, len(headers) + 1):
            cell = ws_resume.cell(row=r, column=c)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if r % 2 == 0:
                cell.fill = PatternFill(start_color=LIGHT_BG, end_color=LIGHT_BG, fill_type='solid')
    
    # Ajuster largeurs colonnes Resume
    ws_resume.column_dimensions['A'].width = 25
    for col_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws_resume.column_dimensions[col_letter].width = 18
    
    # ─────────────────────────────────────────────────────────────────────────
    # FEUILLE 2 : EVOLUTION MENSUELLE
    # ─────────────────────────────────────────────────────────────────────────
    ws_monthly = wb.create_sheet("Evolution Mensuelle")
    ws_monthly.sheet_properties.tabColor = '2563EB'
    
    month_labels = ['Jan', 'Fev', 'Mar', 'Avr', 'Mai', 'Juin', 'Juil', 'Aout', 'Sep', 'Oct', 'Nov', 'Dec']
    
    if not df_transactions.empty and 'datetransaction' in df_transactions.columns:
        df_plot = df_transactions.copy()
        df_plot['year'] = df_plot['datetransaction'].dt.year
        df_plot['month_num'] = df_plot['datetransaction'].dt.month
        
        years = sorted(df_plot['year'].unique(), reverse=True)
        current_yr = years[0] if len(years) > 0 else 2025
        previous_yr = years[1] if len(years) > 1 else current_yr - 1
        
        df_n = df_plot[df_plot['year'] == current_yr]
        df_n1 = df_plot[df_plot['year'] == previous_yr]
        
        # Aggregations mensuelles
        def monthly_sum(df_yr, col):
            if df_yr.empty or col not in df_yr.columns:
                return [0.0] * 12
            grp = df_yr.groupby('month_num')[col].sum()
            return [round(float(grp.get(m, 0.0)), 2) for m in range(1, 13)]
        
        ca_n = monthly_sum(df_n, 'ca')
        vol_gasoil_n = monthly_sum(df_n, 'volume_gasoil')
        vol_super_n = monthly_sum(df_n, 'volume_super')
        marge_n = monthly_sum(df_n, 'marge_ht')
        
        ca_n1 = monthly_sum(df_n1, 'ca')
        vol_gasoil_n1 = monthly_sum(df_n1, 'volume_gasoil')
        vol_super_n1 = monthly_sum(df_n1, 'volume_super')
        
        # ── Tableau CA Mensuel ──
        ws_monthly['A1'].value = "EVOLUTION MENSUELLE DU CHIFFRE D'AFFAIRES"
        ws_monthly['A1'].font = SUBTITLE_FONT
        ws_monthly.merge_cells('A1:N1')
        
        # En-tetes
        r = 3
        headers_m = ['Annee'] + month_labels + ['TOTAL']
        for ci, h in enumerate(headers_m, 1):
            cell = ws_monthly.cell(row=r, column=ci, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal='center')
            cell.border = THIN_BORDER
        
        # Ligne N
        r = 4
        ws_monthly.cell(row=r, column=1, value=str(current_yr)).font = Font(name='Calibri', bold=True, size=10)
        for ci, val in enumerate(ca_n, 2):
            ws_monthly.cell(row=r, column=ci, value=val).number_format = '#,##0'
        ws_monthly.cell(row=r, column=14, value=sum(ca_n)).number_format = '#,##0'
        
        if compare:
            # Ligne N-1
            r = 5
            ws_monthly.cell(row=r, column=1, value=str(previous_yr)).font = Font(name='Calibri', bold=True, size=10)
            for ci, val in enumerate(ca_n1, 2):
                ws_monthly.cell(row=r, column=ci, value=val).number_format = '#,##0'
            ws_monthly.cell(row=r, column=14, value=sum(ca_n1)).number_format = '#,##0'
            
            # Ligne Variation
            r = 6
            ws_monthly.cell(row=r, column=1, value='Variation').font = Font(name='Calibri', bold=True, size=10, color=SALAMA_RED)
            for ci in range(12):
                if ca_n1[ci] > 0:
                    var = (ca_n[ci] - ca_n1[ci]) / ca_n1[ci]
                    var_cell = ws_monthly.cell(row=r, column=ci+2, value=var)
                    var_cell.number_format = '+0.0%;-0.0%'
                    var_cell.font = Font(name='Calibri', bold=True, size=10, 
                                         color='16A34A' if var >= 0 else 'DC2626')
                else:
                    ws_monthly.cell(row=r, column=ci+2, value='N/A')
            last_data_row = 6
        else:
            last_data_row = 4
        
        # Bordures tableau
        for rr in range(3, last_data_row + 1):
            for cc in range(1, 15):
                cell = ws_monthly.cell(row=rr, column=cc)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal='center')
                if rr > 3 and rr % 2 == 0:
                    cell.fill = PatternFill(start_color=LIGHT_BG, end_color=LIGHT_BG, fill_type='solid')
        
        # Graphique CA Mensuel
        chart_ca = BarChart()
        chart_ca.type = "col"
        chart_ca.title = f"Chiffre d'Affaires Mensuel - {client_name}"
        chart_ca.y_axis.title = "CA (MAD)"
        chart_ca.style = 10
        chart_ca.width = 30
        chart_ca.height = 15
        
        cats = Reference(ws_monthly, min_col=2, max_col=13, min_row=3)
        data_n_ref = Reference(ws_monthly, min_col=2, max_col=13, min_row=4)
        chart_ca.add_data(data_n_ref, from_rows=True)
        chart_ca.set_categories(cats)
        chart_ca.series[0].title = SeriesLabel(v=f"CA {current_yr}")
        chart_ca.series[0].graphicalProperties.solidFill = "EAB308"
        
        if compare:
            data_n1_ref = Reference(ws_monthly, min_col=2, max_col=13, min_row=5)
            chart_ca.add_data(data_n1_ref, from_rows=True)
            chart_ca.series[1].title = SeriesLabel(v=f"CA {previous_yr}")
            chart_ca.series[1].graphicalProperties.solidFill = "93C5FD"
        
        ws_monthly.add_chart(chart_ca, f"A{last_data_row + 2}")
        
        # ── Tableau Volumes Mensuels ──
        vol_start = last_data_row + 20
        ws_monthly.cell(row=vol_start, column=1).value = "EVOLUTION MENSUELLE DES VOLUMES"
        ws_monthly.cell(row=vol_start, column=1).font = SUBTITLE_FONT
        ws_monthly.merge_cells(f'A{vol_start}:N{vol_start}')
        
        r = vol_start + 2
        headers_vol = ['Produit'] + month_labels + ['TOTAL']
        for ci, h in enumerate(headers_vol, 1):
            cell = ws_monthly.cell(row=r, column=ci, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal='center')
            cell.border = THIN_BORDER
        
        # Gasoil N
        r += 1
        ws_monthly.cell(row=r, column=1, value=f'Gasoil {current_yr}').font = Font(name='Calibri', bold=True, size=10)
        for ci, val in enumerate(vol_gasoil_n, 2):
            ws_monthly.cell(row=r, column=ci, value=val).number_format = '#,##0'
        ws_monthly.cell(row=r, column=14, value=sum(vol_gasoil_n)).number_format = '#,##0'
        
        # Super N
        r += 1
        ws_monthly.cell(row=r, column=1, value=f'Super SP {current_yr}').font = Font(name='Calibri', bold=True, size=10)
        for ci, val in enumerate(vol_super_n, 2):
            ws_monthly.cell(row=r, column=ci, value=val).number_format = '#,##0'
        ws_monthly.cell(row=r, column=14, value=sum(vol_super_n)).number_format = '#,##0'
        
        if compare:
            r += 1
            ws_monthly.cell(row=r, column=1, value=f'Gasoil {previous_yr}').font = Font(name='Calibri', bold=True, size=10, color=SLATE_600)
            for ci, val in enumerate(vol_gasoil_n1, 2):
                ws_monthly.cell(row=r, column=ci, value=val).number_format = '#,##0'
            ws_monthly.cell(row=r, column=14, value=sum(vol_gasoil_n1)).number_format = '#,##0'
            
            r += 1
            ws_monthly.cell(row=r, column=1, value=f'Super SP {previous_yr}').font = Font(name='Calibri', bold=True, size=10, color=SLATE_600)
            for ci, val in enumerate(vol_super_n1, 2):
                ws_monthly.cell(row=r, column=ci, value=val).number_format = '#,##0'
            ws_monthly.cell(row=r, column=14, value=sum(vol_super_n1)).number_format = '#,##0'
        
        vol_table_end = r
        
        # Bordures tableau volumes
        for rr in range(vol_start + 2, vol_table_end + 1):
            for cc in range(1, 15):
                cell = ws_monthly.cell(row=rr, column=cc)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal='center')
        
        # Graphique Volumes
        chart_vol = BarChart()
        chart_vol.type = "col"
        chart_vol.grouping = "stacked"
        chart_vol.title = f"Volumes Mensuels - {client_name}"
        chart_vol.y_axis.title = "Volume (L)"
        chart_vol.style = 10
        chart_vol.width = 30
        chart_vol.height = 15
        
        cats_vol = Reference(ws_monthly, min_col=2, max_col=13, min_row=vol_start + 2)
        
        # Gasoil N
        gasoil_n_row = vol_start + 3
        data_gasoil_n = Reference(ws_monthly, min_col=2, max_col=13, min_row=gasoil_n_row)
        chart_vol.add_data(data_gasoil_n, from_rows=True)
        chart_vol.set_categories(cats_vol)
        chart_vol.series[0].title = SeriesLabel(v=f"Gasoil {current_yr}")
        chart_vol.series[0].graphicalProperties.solidFill = "EAB308"
        
        # Super N
        super_n_row = vol_start + 4
        data_super_n = Reference(ws_monthly, min_col=2, max_col=13, min_row=super_n_row)
        chart_vol.add_data(data_super_n, from_rows=True)
        chart_vol.series[1].title = SeriesLabel(v=f"Super SP {current_yr}")
        chart_vol.series[1].graphicalProperties.solidFill = "F97316"
        
        ws_monthly.add_chart(chart_vol, f"A{vol_table_end + 2}")
        
        # ── Mix Produit (Pie Chart) ──
        pie_start = vol_table_end + 20
        ws_monthly.cell(row=pie_start, column=1).value = "REPARTITION MIX PRODUIT"
        ws_monthly.cell(row=pie_start, column=1).font = SUBTITLE_FONT
        
        ws_monthly.cell(row=pie_start + 2, column=1, value='Produit').font = HEADER_FONT
        ws_monthly.cell(row=pie_start + 2, column=1).fill = HEADER_FILL
        ws_monthly.cell(row=pie_start + 2, column=1).border = THIN_BORDER
        ws_monthly.cell(row=pie_start + 2, column=2, value=f'Volume {current_yr} (L)').font = HEADER_FONT
        ws_monthly.cell(row=pie_start + 2, column=2).fill = HEADER_FILL
        ws_monthly.cell(row=pie_start + 2, column=2).border = THIN_BORDER
        
        ws_monthly.cell(row=pie_start + 3, column=1, value='Gasoil').border = THIN_BORDER
        ws_monthly.cell(row=pie_start + 3, column=2, value=sum(vol_gasoil_n)).number_format = '#,##0'
        ws_monthly.cell(row=pie_start + 3, column=2).border = THIN_BORDER
        
        ws_monthly.cell(row=pie_start + 4, column=1, value='Super SP').border = THIN_BORDER
        ws_monthly.cell(row=pie_start + 4, column=2, value=sum(vol_super_n)).number_format = '#,##0'
        ws_monthly.cell(row=pie_start + 4, column=2).border = THIN_BORDER
        
        chart_pie = PieChart()
        chart_pie.title = f"Mix Produit {current_yr}"
        chart_pie.style = 10
        chart_pie.width = 14
        chart_pie.height = 12
        
        labels_pie = Reference(ws_monthly, min_col=1, min_row=pie_start + 3, max_row=pie_start + 4)
        data_pie = Reference(ws_monthly, min_col=2, min_row=pie_start + 2, max_row=pie_start + 4)
        chart_pie.add_data(data_pie, titles_from_data=True)
        chart_pie.set_categories(labels_pie)
        
        # Couleurs personnalisees
        s = chart_pie.series[0]
        pt_gasoil = DataPoint(idx=0)
        pt_gasoil.graphicalProperties.solidFill = "EAB308"
        s.data_points.append(pt_gasoil)
        pt_super = DataPoint(idx=1)
        pt_super.graphicalProperties.solidFill = "F97316"
        s.data_points.append(pt_super)
        
        chart_pie.dataLabels = DataLabelList()
        chart_pie.dataLabels.showPercent = True
        chart_pie.dataLabels.showVal = True
        
        ws_monthly.add_chart(chart_pie, f"D{pie_start}")
        
        # Largeurs colonnes
        ws_monthly.column_dimensions['A'].width = 20
        for col_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
            ws_monthly.column_dimensions[col_letter].width = 12
    
    # ─────────────────────────────────────────────────────────────────────────
    # FEUILLE 3 : TOP 10 CLIENTS
    # ─────────────────────────────────────────────────────────────────────────
    ws_top10 = wb.create_sheet("Top 10 Clients")
    ws_top10.sheet_properties.tabColor = '16A34A'
    
    ws_top10['A1'].value = "TOP 10 CLIENTS PAR CHIFFRE D'AFFAIRES"
    ws_top10['A1'].font = SUBTITLE_FONT
    ws_top10.merge_cells('A1:G1')
    
    if not df_transactions.empty and 'client' in df_transactions.columns:
        df_plot = df_transactions.copy()
        df_plot['year'] = df_plot['datetransaction'].dt.year
        years = sorted(df_plot['year'].unique(), reverse=True)
        current_yr = years[0] if years else 2025
        previous_yr = years[1] if len(years) > 1 else current_yr - 1
        
        df_n = df_plot[df_plot['year'] == current_yr]
        df_n1 = df_plot[df_plot['year'] == previous_yr]
        
        agg_cols = {}
        for col in ['ca', 'volume_gasoil', 'volume_super', 'marge_ht']:
            if col in df_n.columns:
                agg_cols[col] = 'sum'
        
        if agg_cols and not df_n.empty:
            top_n = df_n.groupby('client').agg(agg_cols).nlargest(10, 'ca' if 'ca' in agg_cols else list(agg_cols.keys())[0]).reset_index()
            
            ca_n1_dict = {}
            if not df_n1.empty and 'ca' in df_n1.columns:
                ca_n1_dict = df_n1.groupby('client')['ca'].sum().to_dict()
            
            # En-tetes
            r = 3
            if compare:
                headers_t = ['Client', f'CA {current_yr} (MAD)', f'CA {previous_yr} (MAD)', 'Variation', 'Vol. Gasoil (L)', 'Vol. Super (L)', 'Marge HT (MAD)']
            else:
                headers_t = ['Client', f'CA {current_yr} (MAD)', 'Vol. Gasoil (L)', 'Vol. Super (L)', 'Marge HT (MAD)']
            
            for ci, h in enumerate(headers_t, 1):
                cell = ws_top10.cell(row=r, column=ci, value=h)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = Alignment(horizontal='center')
                cell.border = THIN_BORDER
            
            for i, (_, row_data) in enumerate(top_n.iterrows()):
                rr = r + 1 + i
                client = str(row_data['client'])
                ca = float(row_data.get('ca', 0))
                ca_prev = float(ca_n1_dict.get(client, 0))
                vol_g = float(row_data.get('volume_gasoil', 0))
                vol_s = float(row_data.get('volume_super', 0))
                marge = float(row_data.get('marge_ht', 0))
                
                if compare:
                    ws_top10.cell(row=rr, column=1, value=client)
                    ws_top10.cell(row=rr, column=2, value=round(ca, 2)).number_format = '#,##0.00'
                    ws_top10.cell(row=rr, column=3, value=round(ca_prev, 2)).number_format = '#,##0.00'
                    
                    var_cell = ws_top10.cell(row=rr, column=4)
                    if ca_prev > 0:
                        var = (ca - ca_prev) / ca_prev
                        var_cell.value = var
                        var_cell.number_format = '+0.0%;-0.0%'
                        var_cell.font = Font(name='Calibri', bold=True, color='16A34A' if var >= 0 else 'DC2626')
                    else:
                        var_cell.value = 'N/A'
                    
                    ws_top10.cell(row=rr, column=5, value=round(vol_g, 0)).number_format = '#,##0'
                    ws_top10.cell(row=rr, column=6, value=round(vol_s, 0)).number_format = '#,##0'
                    ws_top10.cell(row=rr, column=7, value=round(marge, 2)).number_format = '#,##0.00'
                else:
                    ws_top10.cell(row=rr, column=1, value=client)
                    ws_top10.cell(row=rr, column=2, value=round(ca, 2)).number_format = '#,##0.00'
                    ws_top10.cell(row=rr, column=3, value=round(vol_g, 0)).number_format = '#,##0'
                    ws_top10.cell(row=rr, column=4, value=round(vol_s, 0)).number_format = '#,##0'
                    ws_top10.cell(row=rr, column=5, value=round(marge, 2)).number_format = '#,##0.00'
                
                # Style alternance
                for cc in range(1, len(headers_t) + 1):
                    cell = ws_top10.cell(row=rr, column=cc)
                    cell.border = THIN_BORDER
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    if rr % 2 == 0:
                        cell.fill = PatternFill(start_color=LIGHT_BG, end_color=LIGHT_BG, fill_type='solid')
            
            top10_end = r + len(top_n) + 1
            
            # Graphique Top 10
            chart_top = BarChart()
            chart_top.type = "bar"
            chart_top.title = f"Top 10 Clients par CA - {current_yr}"
            chart_top.x_axis.title = "CA (MAD)"
            chart_top.style = 10
            chart_top.width = 25
            chart_top.height = 14
            
            ca_col = 2
            cats_top = Reference(ws_top10, min_col=1, min_row=r + 1, max_row=top10_end - 1)
            data_top = Reference(ws_top10, min_col=ca_col, min_row=r, max_row=top10_end - 1)
            chart_top.add_data(data_top, titles_from_data=True)
            chart_top.set_categories(cats_top)
            chart_top.series[0].graphicalProperties.solidFill = "2563EB"
            
            ws_top10.add_chart(chart_top, f"A{top10_end + 1}")
        
        # Largeurs
        ws_top10.column_dimensions['A'].width = 30
        for col_letter in ['B', 'C', 'D', 'E', 'F', 'G']:
            ws_top10.column_dimensions[col_letter].width = 18
    
    # ─────────────────────────────────────────────────────────────────────────
    # FEUILLE 4 : DETAIL TRANSACTIONS
    # ─────────────────────────────────────────────────────────────────────────
    ws_detail = wb.create_sheet("Detail Transactions")
    ws_detail.sheet_properties.tabColor = 'F59E0B'
    
    if not df_transactions.empty:
        # Selectionner les colonnes pertinentes
        detail_cols_map = {
            'datetransaction': 'Date',
            'client': 'Client',
            'statut': 'Statut',
            'volume_gasoil': 'Vol. Gasoil (L)',
            'volume_super': 'Vol. Super (L)',
            'ca': 'CA (MAD)',
            'marge_ht': 'Marge HT (MAD)',
            'fournisseur': 'Fournisseur',
            'prix_vente_gasoil_ttc': 'PV Gasoil TTC',
            'prix_vente_super_ttc': 'PV Super TTC',
        }
        
        available_cols = [c for c in detail_cols_map.keys() if c in df_transactions.columns]
        df_detail = df_transactions[available_cols].copy()
        
        # Formater dates
        if 'datetransaction' in df_detail.columns:
            df_detail['datetransaction'] = df_detail['datetransaction'].dt.strftime('%d/%m/%Y')
        
        # En-tetes
        r = 1
        for ci, col in enumerate(available_cols, 1):
            cell = ws_detail.cell(row=r, column=ci, value=detail_cols_map[col])
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal='center')
            cell.border = THIN_BORDER
        
        # Donnees
        for ri, (_, row_data) in enumerate(df_detail.iterrows()):
            for ci, col in enumerate(available_cols, 1):
                val = row_data[col]
                cell = ws_detail.cell(row=r + 1 + ri, column=ci, value=val)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal='center')
                cell.font = NORMAL_FONT
                if (r + 1 + ri) % 2 == 0:
                    cell.fill = PatternFill(start_color=LIGHT_BG, end_color=LIGHT_BG, fill_type='solid')
                
                # Format nombres
                if col in ['volume_gasoil', 'volume_super']:
                    cell.number_format = '#,##0'
                elif col in ['ca', 'marge_ht', 'prix_vente_gasoil_ttc', 'prix_vente_super_ttc']:
                    cell.number_format = '#,##0.00'
        
        # Largeurs auto
        for ci, col in enumerate(available_cols, 1):
            ws_detail.column_dimensions[get_column_letter(ci)].width = max(len(detail_cols_map[col]) + 4, 14)
    
    # ── Sauvegarder en bytes ──────────────────────────────────────────────────
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
